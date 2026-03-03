import sys
import re
import openpyxl
import matplotlib.pyplot as plt


# =========================
# КОНСТАНТЫ
# =========================
DEFAULT_XLSX = "openmp_scalability_with_chart.xlsx"  # поставь имя/путь по умолчанию
SHEET_NAME = "Table"
OUT_PNG = "speedup.png"
OUT_PDF = "speedup.pdf"

P_LIST = [1, 2, 4, 7, 8, 16, 20, 40]


def as_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(" ", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def extract_size(v):
    """
    Возвращает int 20000/40000 из:
    - числа (20000)
    - строки ("20000 (~ 3 GiB)", "40 000", "40000\n(~ 12 GiB)" и т.п.)
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)

    if isinstance(v, str):
        # убираем пробелы, но оставляем цифры
        s = v.replace(" ", "")
        m = re.search(r"(20000|40000)", s)
        if m:
            return int(m.group(1))
    return None


def find_sheet(wb):
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    return wb.active


def find_header_row_with_T1(ws, max_rows=40, max_cols=100):
    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "T1":
                return r
    raise RuntimeError("Не нашёл заголовок 'T1'. Проверь лист/шаблон.")


def build_col_map(ws, header_row, max_cols=200):
    col_map = {}
    for c in range(1, max_cols + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str):
            key = v.strip().upper()
            if key.startswith("T") or key.startswith("S"):
                col_map[key] = c
    return col_map


def find_size_column(ws, header_row, sizes=(20000, 40000), max_rows=120, max_cols=30):
    need = set(sizes)
    for c in range(1, max_cols + 1):
        found = set()
        for r in range(header_row + 1, max_rows + 1):
            v = ws.cell(r, c).value
            s = extract_size(v)
            if s in need:
                found.add(s)
        if found == need:
            return c
    raise RuntimeError("Не нашёл колонку с 20000 и 40000. Скорее всего размеры записаны нестандартно.")


def find_row_for_size(ws, size_col, size_value, start_row, max_rows=200):
    for r in range(start_row, max_rows + 1):
        v = ws.cell(r, size_col).value
        s = extract_size(v)
        if s == size_value:
            return r
    raise RuntimeError(f"Не нашёл строку для размера {size_value}.")


def read_times(ws, row, col_map):
    times = {}
    for p in P_LIST:
        key = f"T{p}"
        if key not in col_map:
            raise RuntimeError(f"Не найден столбец '{key}' в заголовках.")
        v = ws.cell(row, col_map[key]).value
        num = as_number(v)
        if num is None:
            raise RuntimeError(f"Пустое/нечисловое значение {key} в строке {row}.")
        times[p] = num
    return times


def speedups(times):
    t1 = times[1]
    return {p: (t1 / times[p]) for p in times}


def main():
    xlsx = DEFAULT_XLSX
    if len(sys.argv) >= 2:
        xlsx = sys.argv[1]

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = find_sheet(wb)

    header_row = find_header_row_with_T1(ws)
    col_map = build_col_map(ws, header_row)

    size_col = find_size_column(ws, header_row, sizes=(20000, 40000))
    row_20000 = find_row_for_size(ws, size_col, 20000, header_row + 1)
    row_40000 = find_row_for_size(ws, size_col, 40000, header_row + 1)

    t20000 = read_times(ws, row_20000, col_map)
    t40000 = read_times(ws, row_40000, col_map)

    s20000 = speedups(t20000)
    s40000 = speedups(t40000)

    p = P_LIST
    y_lin = p
    y_20000 = [s20000[x] for x in p]
    y_40000 = [s40000[x] for x in p]

    plt.plot(p, y_lin, "--o", label="Linear (S=p)")
    plt.plot(p, y_20000, "-o", label="N=20000")
    plt.plot(p, y_40000, "-o", label="N=40000")

    plt.xlabel("p (threads)")
    plt.ylabel("S(p) = T1 / Tp")
    plt.title("OpenMP scalability")
    plt.grid(True)
    plt.legend()

    plt.savefig(OUT_PNG, dpi=200)
    plt.savefig(OUT_PDF)
    print(f"OK: saved {OUT_PNG} and {OUT_PDF}")


if __name__ == "__main__":
    main()